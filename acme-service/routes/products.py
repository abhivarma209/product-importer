from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

from dependencies.database import get_db
from models import Product
from schemas import ProductCreate, ProductUpdate, ProductResponse
from app.tasks import trigger_webhooks_async

router = APIRouter(prefix="/api/products", tags=["products"])


@router.get("", response_model=dict)
def get_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=1000),
    search: Optional[str] = None,
    active: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Get paginated list of products with filtering"""
    query = db.query(Product)
    
    # Apply filters
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                func.lower(Product.sku).like(func.lower(search_filter)),
                func.lower(Product.name).like(func.lower(search_filter)),
                func.lower(Product.description).like(func.lower(search_filter))
            )
        )
    
    if active is not None:
        query = query.filter(Product.active == active)
    
    # Get total count
    total = query.count()
    
    # Get paginated results
    products = query.order_by(Product.id.desc()).offset(skip).limit(limit).all()
    
    return {
        "items": products,
        "total": total,
        "skip": skip,
        "limit": limit
    }


@router.post("", response_model=ProductResponse, status_code=201)
def create_product(product: ProductCreate, db: Session = Depends(get_db)):
    """Create a new product"""
    # Check if SKU already exists (case-insensitive)
    existing = db.query(Product).filter(
        func.lower(Product.sku) == func.lower(product.sku)
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")
    
    db_product = Product(**product.model_dump())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    
    # Trigger webhooks
    trigger_webhooks_async.delay('product.created', {'product_id': db_product.id, 'sku': db_product.sku})
    
    return db_product


@router.get("/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    """Get a single product by ID"""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: int,
    product_update: ProductUpdate,
    db: Session = Depends(get_db)
):
    """Update a product"""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check SKU uniqueness if updating SKU
    if product_update.sku and product_update.sku != product.sku:
        existing = db.query(Product).filter(
            func.lower(Product.sku) == func.lower(product_update.sku)
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="SKU already exists")
    
    # Update fields
    update_data = product_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(product, field, value)
    
    db.commit()
    db.refresh(product)
    
    # Trigger webhooks
    trigger_webhooks_async.delay('product.updated', {'product_id': product.id, 'sku': product.sku})
    
    return product


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    """Delete a single product"""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    sku = product.sku
    db.delete(product)
    db.commit()
    
    # Trigger webhooks
    trigger_webhooks_async.delay('product.deleted', {'product_id': product_id, 'sku': sku})
    
    return {"message": "Product deleted successfully"}


@router.delete("")
def bulk_delete_products(db: Session = Depends(get_db)):
    """Delete all products"""
    count = db.query(Product).count()
    db.query(Product).delete()
    db.commit()
    
    # Trigger webhooks
    trigger_webhooks_async.delay('products.bulk_deleted', {'count': count})
    
    return {"message": f"Successfully deleted {count} products", "count": count}


@router.get("/export/excel")
def export_products_to_excel(
    search: Optional[str] = None,
    active: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Export products to Excel file"""
    
    # Query products with filters
    query = db.query(Product)
    
    # Apply filters
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                func.lower(Product.sku).like(func.lower(search_filter)),
                func.lower(Product.name).like(func.lower(search_filter)),
                func.lower(Product.description).like(func.lower(search_filter))
            )
        )
    
    if active is not None:
        query = query.filter(Product.active == active)
    
    # Get all products
    products = query.order_by(Product.id).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Define headers
    headers = ["ID", "SKU", "Name", "Description", "Price", "Status", "Created At"]
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.id)
        ws.cell(row=row_num, column=2, value=product.sku)
        ws.cell(row=row_num, column=3, value=product.name)
        ws.cell(row=row_num, column=4, value=product.description or "")
        ws.cell(row=row_num, column=5, value=product.price or 0.0)
        ws.cell(row=row_num, column=6, value="Active" if product.active else "Inactive")
        ws.cell(row=row_num, column=7, value=product.created_at.strftime("%Y-%m-%d %H:%M:%S") if product.created_at else "")
    
    # Auto-adjust column widths
    column_widths = {
        'A': 10,
        'B': 15,
        'C': 30,
        'D': 40,
        'E': 12,
        'F': 12,
        'G': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with timestamp
    filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
